#!/usr/bin/env python3
"""
Excel 内容对比工具（忽略格式，支持插入/删除行检测）
安装依赖: pip install openpyxl
用法:     python cmpexcel.py 表1.xlsx 表2.xlsx
"""

import sys
import difflib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read_sheet_rows(ws):
    """读取sheet所有行，每行为 list of str"""
    rows = []
    for r in range(1, (ws.max_row or 0) + 1):
        row = []
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=r, column=c).value
            row.append("" if v is None else str(v))
        rows.append(row)
    return rows


def format_row(row_idx, row_data):
    """把一行数据格式化成  第N行  A1['x']  B1['y'] ..."""
    parts = []
    for c_idx, val in enumerate(row_data):
        ref = f"{get_column_letter(c_idx + 1)}{row_idx}"
        parts.append(f"{ref}['{val}']")
    return f"  第{row_idx}行  " + "  ".join(parts)


def compare_sheets(ws1, ws2, file1, file2, sheet_label):
    rows1 = read_sheet_rows(ws1)
    rows2 = read_sheet_rows(ws2)

    # 把每行转成字符串用于 diff 匹配
    str_rows1 = ["\t".join(r) for r in rows1]
    str_rows2 = ["\t".join(r) for r in rows2]

    matcher = difflib.SequenceMatcher(None, str_rows1, str_rows2, autojunk=False)
    opcodes = matcher.get_opcodes()

    output = []
    has_diff = False

    for tag, i1, i2, j1, j2 in opcodes:
        if tag == "equal":
            continue

        has_diff = True

        if tag == "replace":
            # 逐行对比，只显示有差异的列
            for di, (r1_idx, r2_idx) in enumerate(zip(range(i1, i2), range(j1, j2))):
                row1 = rows1[r1_idx]
                row2 = rows2[r2_idx]
                max_cols = max(len(row1), len(row2))
                diff_cols = []
                for c in range(max_cols):
                    v1 = row1[c] if c < len(row1) else ""
                    v2 = row2[c] if c < len(row2) else ""
                    if v1 != v2:
                        diff_cols.append(c)
                if diff_cols:
                    output.append(f"  差异 第{r1_idx+1}行")
                    output.append(f"  {file1} Sheet: {sheet_label}")
                    parts1 = f"  第{r1_idx+1}行  " + "  ".join(
                        f"{get_column_letter(c+1)}{r1_idx+1}['{row1[c] if c < len(row1) else ''}']"
                        for c in diff_cols
                    )
                    output.append(parts1)
                    output.append(f"  {file2} Sheet: {sheet_label}")
                    parts2 = f"  第{r2_idx+1}行  " + "  ".join(
                        f"{get_column_letter(c+1)}{r2_idx+1}['{row2[c] if c < len(row2) else ''}']"
                        for c in diff_cols
                    )
                    output.append(parts2)
                    output.append("")

            # replace 块里 行数不等的部分（多出来的行）
            extra1 = list(range(i1 + (j2-j1), i2))  # file1 多出的行
            extra2 = list(range(j1 + (i2-i1), j2))  # file2 多出的行
            for r in extra1:
                output.append(f"  仅 {file1} 有 第{r+1}行")
                output.append(format_row(r+1, rows1[r]))
                output.append("")
            for r in extra2:
                output.append(f"  仅 {file2} 有 第{r+1}行")
                output.append(format_row(r+1, rows2[r]))
                output.append("")

        elif tag == "insert":
            for r in range(j1, j2):
                output.append(f"  仅 {file2} 有 第{r+1}行")
                output.append(format_row(r+1, rows2[r]))
                output.append("")

        elif tag == "delete":
            for r in range(i1, i2):
                output.append(f"  仅 {file1} 有 第{r+1}行")
                output.append(format_row(r+1, rows1[r]))
                output.append("")

    return has_diff, output


def main():
    if len(sys.argv) != 3:
        print("用法: python cmpexcel.py 表1.xlsx 表2.xlsx")
        sys.exit(1)

    file1, file2 = sys.argv[1], sys.argv[2]

    try:
        wb1 = load_workbook(file1, data_only=True)
    except FileNotFoundError:
        print(f"❌ 找不到文件: {file1}"); sys.exit(1)

    try:
        wb2 = load_workbook(file2, data_only=True)
    except FileNotFoundError:
        print(f"❌ 找不到文件: {file2}"); sys.exit(1)

    sheets1, sheets2 = wb1.sheetnames, wb2.sheetnames

    print(f"\n{'='*60}")
    print(f"  {file1}")
    print(f"  {file2}")
    print(f"{'='*60}")

    only_s1 = [s for s in sheets1 if s not in sheets2]
    only_s2 = [s for s in sheets2 if s not in sheets1]
    if only_s1:
        print(f"  ⚠️  仅 {file1} 有的 Sheet: {only_s1}")
    if only_s2:
        print(f"  ⚠️  仅 {file2} 有的 Sheet: {only_s2}")

    total_diff_sheets = 0

    for s1, s2 in zip(sheets1, sheets2):
        label = s1 if s1 == s2 else f"{s1} vs {s2}"
        has_diff, output = compare_sheets(wb1[s1], wb2[s2], file1, file2, label)

        print(f"\n{'─'*60}")
        print(f"  📄 Sheet: {label}")
        print(f"{'─'*60}")

        if not has_diff:
            print("  ✅ 内容完全一致")
        else:
            total_diff_sheets += 1
            for line in output:
                print(line)

    print(f"{'='*60}")
    print(f"  {'🎉 所有 Sheet 内容一致！' if total_diff_sheets == 0 else f'共 {total_diff_sheets} 个 Sheet 有差异'}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
