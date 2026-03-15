#!/usr/bin/env python3
"""
write_excel.py — 向 Excel 文件写入数据，专为 AI 工具调用设计。

操作前自动备份，原子写入，绝不破坏原有格式和公式。

━━━ 操作模式（--mode）━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  set-cell    写入一个或多个指定单元格的值
  append      在数据末尾追加行
  insert      在指定行前插入行
  replace     替换指定行范围的内容
  delete      删除指定行（整行）
  patch       全局查找替换（字符串或正则，跨所有单元格）
  set-col     设置某列所有数据行的值（公式或固定值）
  add-sheet   新增工作表
  del-sheet   删除工作表
  rename-sheet 重命名工作表

━━━ 数据输入格式 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  --data 支持三种形式：
    JSON 数组（对象列表）：[{"姓名":"张三","分数":90}, ...]
    JSON 二维数组：          [["张三", 90], ["李四", 85]]
    CSV 格式字符串：         张三,90\n李四,85

  --data-file  从 UTF-8 JSON 或 CSV 文件读取数据

━━━ 安全特性 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  • 写入前自动生成 .bak 备份（--no-backup 可关闭）
  • 原子写入：先写临时文件，再 rename
  • 保留所有未修改单元格的格式、公式、批注
  • --dry-run 预览操作不写入

━━━ 典型用法 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  # 写入单元格
  python write_excel.py file.xlsx --mode set-cell --cells A1="标题" B2=100 C3="=SUM(B1:B2)"

  # 追加数据行（JSON）
  python write_excel.py file.xlsx --mode append \\
      --data '[{"姓名":"王五","分数":95}]'

  # 在第 3 行前插入
  python write_excel.py file.xlsx --mode insert --start 3 \\
      --data '[["新行数据", 100]]'

  # 替换第 5~7 行
  python write_excel.py file.xlsx --mode replace --start 5 --end 7 \\
      --data-file new_rows.json

  # 删除第 4 行
  python write_excel.py file.xlsx --mode delete --start 4 --end 4

  # 全局替换文本
  python write_excel.py file.xlsx --mode patch --old "旧文本" --new "新文本"
  python write_excel.py file.xlsx --mode patch --old "\\bv\\d+" --new "v2.0" --regex

  # 给 D 列所有数据行写入公式
  python write_excel.py file.xlsx --mode set-col --col D --value "=B{row}*C{row}"

  # 新增工作表
  python write_excel.py file.xlsx --mode add-sheet --sheet-name "汇总"

  # 删除工作表
  python write_excel.py file.xlsx --mode del-sheet --sheet "Sheet3"

  # 重命名工作表
  python write_excel.py file.xlsx --mode rename-sheet --sheet "Sheet1" --sheet-name "数据"
"""

from __future__ import annotations
import sys
import os
import re
import json
import csv
import io
import shutil
import tempfile
import argparse
import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:
    print("错误：请先安装 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════════

def parse_cell_ref(ref: str) -> tuple[int, int]:
    ref = ref.strip().upper()
    col_str = re.match(r"[A-Z]+", ref)
    row_str = re.search(r"\d+", ref)
    if not col_str or not row_str:
        _err(f"无效的单元格引用：{ref}")
    return int(row_str.group()), column_index_from_string(col_str.group())


def col_letter(n: int) -> str:
    return get_column_letter(n)


def backup(path: str) -> str | None:
    if not os.path.exists(path):
        return None
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = f"{path}.{ts}.bak"
    shutil.copy2(path, bak)
    return bak


def atomic_save(wb, path: str) -> None:
    """原子写入：写临时文件后 rename，避免中途失败损坏原文件。"""
    dir_name = os.path.dirname(os.path.abspath(path)) or "."
    tmp_fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".xlsx.tmp")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def load_or_create(path: str, create: bool = False):
    """加载工作簿；不存在时若 create=True 则新建。"""
    if os.path.exists(path):
        try:
            # keep_vba=True 保留宏；data_only=False 保留公式字符串
            return load_workbook(path, keep_vba=True, data_only=False)
        except Exception as e:
            _err(f"无法打开文件：{e}")
    if create:
        print(f"[write_excel] 新建文件：{path}", file=sys.stderr)
        return Workbook()
    _err(f"文件不存在：{path}\n提示：使用 --create 可自动创建。")


def get_sheet(wb, sheet_arg: str | None):
    if sheet_arg is None:
        return wb.active or wb.worksheets[0]
    if sheet_arg in wb.sheetnames:
        return wb[sheet_arg]
    try:
        return wb.worksheets[int(sheet_arg)]
    except (ValueError, IndexError):
        pass
    _err(f"工作表 '{sheet_arg}' 不存在。可用：{wb.sheetnames}")


def find_data_end_row(ws) -> int:
    """找到最后一个非空行的行号（1-indexed），用于确定追加位置。"""
    last = 0
    for row in ws.iter_rows():
        if any(c.value is not None for c in row):
            last = row[0].row
    return last


def parse_data(data_str: str | None, data_file: str | None) -> list[list[Any]] | None:
    """
    解析用户输入的数据，统一返回二维列表 [[col1, col2, ...], ...]。
    支持：JSON 对象数组、JSON 二维数组、CSV 字符串。
    """
    if data_str is None and data_file is None:
        return None

    raw = data_str
    if data_file:
        with open(data_file, "r", encoding="utf-8") as f:
            raw = f.read()

    raw = raw.strip()

    # 尝试 JSON
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if not parsed:
                return []
            if isinstance(parsed[0], dict):
                # 对象数组：取所有键作为列顺序
                keys = list(parsed[0].keys())
                return [[row.get(k, "") for k in keys] for row in parsed]
            if isinstance(parsed[0], list):
                return parsed
            # 一维数组 → 单行
            return [parsed]
        except json.JSONDecodeError:
            pass

    # 尝试 CSV
    reader = csv.reader(io.StringIO(raw))
    rows = [r for r in reader if any(c.strip() for c in r)]
    return rows if rows else []


def coerce_value(v: Any) -> Any:
    """尝试将字符串转换为数字；以 = 开头的保留为公式字符串。"""
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("="):
            return s  # 公式
        try:
            return int(s)
        except ValueError:
            pass
        try:
            return float(s)
        except ValueError:
            pass
    return v


# ══════════════════════════════════════════════════════════════════════════════
# 各操作模式
# ══════════════════════════════════════════════════════════════════════════════

def op_set_cell(ws, cell_assignments: list[str]) -> int:
    """
    写入指定单元格。
    cell_assignments 格式：["A1=值", "B2=100", "C3==SUM(A1:A10)"]
    公式写法：C3==SUM(...) 或 C3="=SUM(...)"（第一个 = 是分隔符）
    返回写入的单元格数。
    """
    count = 0
    for item in cell_assignments:
        if "=" not in item:
            _err(f"--cells 格式错误，应为 单元格=值，例如 A1=100，当前：{item!r}")
        addr, _, val_str = item.partition("=")
        addr = addr.strip().upper()
        # 尝试解析单元格地址
        r, c = parse_cell_ref(addr)
        ws.cell(row=r, column=c, value=coerce_value(val_str))
        count += 1
    return count


def op_append(ws, rows: list[list[Any]]) -> int:
    """在数据末尾追加行，返回追加的行数。"""
    last = find_data_end_row(ws)
    start_row = last + 1
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=start_row + i, column=j + 1, value=coerce_value(val))
    return len(rows)


def op_insert(ws, rows: list[list[Any]], before_row: int) -> int:
    """
    在 before_row（1-indexed Excel 行）前插入若干行。
    先用 openpyxl 的 insert_rows 移位，再填数据。
    """
    n = len(rows)
    ws.insert_rows(before_row, amount=n)
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=before_row + i, column=j + 1, value=coerce_value(val))
    return n


def op_replace(ws, rows: list[list[Any]], start_row: int, end_row: int) -> tuple[int, int]:
    """
    替换 [start_row, end_row] 行区间（Excel 行号）。
    先删除原有行，再插入新行。
    返回 (删除行数, 插入行数)。
    """
    old_count = end_row - start_row + 1
    ws.delete_rows(start_row, old_count)
    n_new = len(rows)
    if n_new > 0:
        ws.insert_rows(start_row, amount=n_new)
        for i, row_data in enumerate(rows):
            for j, val in enumerate(row_data):
                ws.cell(row=start_row + i, column=j + 1, value=coerce_value(val))
    return old_count, n_new


def op_delete(ws, start_row: int, end_row: int) -> int:
    """删除 [start_row, end_row] 行，返回删除行数。"""
    count = end_row - start_row + 1
    ws.delete_rows(start_row, count)
    return count


def op_patch(
    ws,
    old: str,
    new: str,
    *,
    use_regex: bool = False,
    count: int = 0,
    sheet_scope: bool = False,  # 保留：单表 scope（默认单表）
) -> int:
    """
    遍历工作表所有单元格，对字符串值进行查找替换。
    公式单元格（以 = 开头）也会被替换（替换公式文本）。
    返回替换次数。
    """
    flags = re.IGNORECASE
    compiled = re.compile(old, flags) if use_regex else None
    total_replacements = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val_str = str(cell.value)
            if use_regex:
                new_val, n = re.subn(old, new, val_str, count=count, flags=flags)
            else:
                if old not in val_str:
                    continue
                if count:
                    new_val = val_str.replace(old, new, count)
                    n = min(val_str.count(old), count)
                else:
                    new_val = val_str.replace(old, new)
                    n = val_str.count(old)
            if n > 0:
                # 保持类型
                if isinstance(cell.value, (int, float)):
                    try:
                        cell.value = type(cell.value)(new_val)
                    except (ValueError, TypeError):
                        cell.value = new_val
                else:
                    cell.value = coerce_value(new_val)
                total_replacements += n
    return total_replacements


def op_set_col(ws, col: str, value_template: str, has_header: bool = True) -> int:
    """
    对指定列所有数据行写入固定值或模板化公式。
    value_template 中 {row} 会被替换为实际 Excel 行号。
    返回写入的单元格数。
    """
    try:
        col_idx = column_index_from_string(col.upper())
    except Exception:
        _err(f"无效的列标识：{col}")

    # 确定数据行范围
    max_row = find_data_end_row(ws)
    start = 2 if has_header else 1
    count = 0
    for r in range(start, max_row + 1):
        val = value_template.replace("{row}", str(r))
        ws.cell(row=r, column=col_idx, value=coerce_value(val))
        count += 1
    return count


def op_add_sheet(wb, name: str, position: int | None = None) -> None:
    if name in wb.sheetnames:
        _err(f"工作表 '{name}' 已存在。")
    wb.create_sheet(title=name, index=position)


def op_del_sheet(wb, sheet_arg: str) -> str:
    ws = get_sheet(wb, sheet_arg)
    name = ws.title
    del wb[name]
    return name


def op_rename_sheet(wb, sheet_arg: str, new_name: str) -> str:
    ws = get_sheet(wb, sheet_arg)
    old_name = ws.title
    ws.title = new_name
    return old_name


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

MODES = (
    "set-cell", "append", "insert", "replace", "delete",
    "patch", "set-col", "add-sheet", "del-sheet", "rename-sheet",
)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="向 Excel 文件写入数据，专为 AI 工具调用设计。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    # ── 必填 ────────────────────────────────────────────────────────────────
    p.add_argument("filename", help="目标 Excel 文件路径（.xlsx）")
    p.add_argument("--mode", choices=MODES, required=True,
                   help=f"操作模式")

    # ── 工作表定位 ───────────────────────────────────────────────────────────
    p.add_argument("--sheet", default=None, metavar="NAME_OR_INDEX",
                   help="工作表名称或索引（0-based），默认第一个")

    # ── 单元格写入 ───────────────────────────────────────────────────────────
    p.add_argument("--cells", nargs="+", default=None, metavar="ADDR=VAL",
                   help='[set-cell] 如 A1=100 B2=标题 C3="=SUM(A1:A10)"')

    # ── 数据输入 ─────────────────────────────────────────────────────────────
    data_src = p.add_mutually_exclusive_group()
    data_src.add_argument("--data", default=None, metavar="JSON_OR_CSV",
                          help="JSON 数组或 CSV 格式数据")
    data_src.add_argument("--data-file", default=None, metavar="FILE",
                          help="从 UTF-8 JSON/CSV 文件读取数据")

    # ── 行范围 ───────────────────────────────────────────────────────────────
    p.add_argument("--start", type=int, default=None, metavar="N",
                   help="起始 Excel 行号（1-indexed，含表头）")
    p.add_argument("--end", type=int, default=None, metavar="N",
                   help="结束 Excel 行号（含）")

    # ── patch 专用 ───────────────────────────────────────────────────────────
    p.add_argument("--old", default=None, metavar="PATTERN",
                   help="[patch] 要替换的旧文本或正则模式")
    p.add_argument("--new", default=None, metavar="TEXT",
                   help="[patch] 替换后的新文本")
    p.add_argument("--regex", action="store_true",
                   help="[patch] 将 --old 视为正则表达式")
    p.add_argument("--count", type=int, default=0, metavar="N",
                   help="[patch] 每格最多替换 N 次（0 = 全部）")

    # ── set-col 专用 ─────────────────────────────────────────────────────────
    p.add_argument("--col", default=None, metavar="COL",
                   help="[set-col] 列字母，如 D")
    p.add_argument("--value", default=None, metavar="VAL_OR_FORMULA",
                   help='[set-col] 值或公式模板，{row} 会被替换为行号，如 "=B{row}*C{row}"')
    p.add_argument("--no-header", action="store_true",
                   help="[set-col] 无表头，从第 1 行开始写")

    # ── 工作表操作 ───────────────────────────────────────────────────────────
    p.add_argument("--sheet-name", default=None, metavar="NAME",
                   help="[add-sheet/rename-sheet] 新工作表名称")
    p.add_argument("--sheet-position", type=int, default=None, metavar="N",
                   help="[add-sheet] 插入位置（0-based），默认末尾")

    # ── 安全 / 预览 ──────────────────────────────────────────────────────────
    p.add_argument("--no-backup", action="store_true", help="不生成 .bak 备份")
    p.add_argument("--dry-run", action="store_true", help="预览，不实际写入文件")
    p.add_argument("--create", action="store_true", help="文件不存在时自动创建")

    return p


def main() -> None:  # noqa: C901
    args = build_parser().parse_args()
    path = os.path.abspath(args.filename)
    mode = args.mode

    # ── 加载工作簿 ────────────────────────────────────────────────────────────
    wb = load_or_create(path, create=args.create)

    extra_info = ""

    try:
        # ── set-cell ──────────────────────────────────────────────────────────
        if mode == "set-cell":
            if not args.cells:
                _err("--mode set-cell 需要 --cells，例如 --cells A1=100 B2=标题")
            ws = get_sheet(wb, args.sheet)
            n = op_set_cell(ws, args.cells)
            extra_info = f"写入 {n} 个单元格"

        # ── append ────────────────────────────────────────────────────────────
        elif mode == "append":
            rows = parse_data(args.data, args.data_file)
            if rows is None:
                _err("--mode append 需要 --data 或 --data-file")
            ws = get_sheet(wb, args.sheet)
            n = op_append(ws, rows)
            extra_info = f"追加 {n} 行"

        # ── insert ────────────────────────────────────────────────────────────
        elif mode == "insert":
            rows = parse_data(args.data, args.data_file)
            if rows is None:
                _err("--mode insert 需要 --data 或 --data-file")
            if args.start is None:
                _err("--mode insert 需要 --start（要插入到哪行之前）")
            ws = get_sheet(wb, args.sheet)
            n = op_insert(ws, rows, args.start)
            extra_info = f"在第 {args.start} 行前插入 {n} 行"

        # ── replace ───────────────────────────────────────────────────────────
        elif mode == "replace":
            rows = parse_data(args.data, args.data_file)
            if rows is None:
                _err("--mode replace 需要 --data 或 --data-file")
            if args.start is None:
                _err("--mode replace 需要 --start")
            end = args.end if args.end is not None else args.start
            ws = get_sheet(wb, args.sheet)
            old_n, new_n = op_replace(ws, rows, args.start, end)
            extra_info = f"替换第 {args.start}～{end} 行（删除 {old_n} 行，插入 {new_n} 行）"

        # ── delete ────────────────────────────────────────────────────────────
        elif mode == "delete":
            if args.start is None:
                _err("--mode delete 需要 --start")
            end = args.end if args.end is not None else args.start
            ws = get_sheet(wb, args.sheet)
            n = op_delete(ws, args.start, end)
            extra_info = f"删除第 {args.start}～{end} 行（共 {n} 行）"

        # ── patch ─────────────────────────────────────────────────────────────
        elif mode == "patch":
            if args.old is None:
                _err("--mode patch 需要 --old")
            new_val = args.new if args.new is not None else ""
            ws = get_sheet(wb, args.sheet)
            n = op_patch(ws, args.old, new_val, use_regex=args.regex, count=args.count)
            extra_info = f"替换 {n} 处"

        # ── set-col ───────────────────────────────────────────────────────────
        elif mode == "set-col":
            if not args.col:
                _err("--mode set-col 需要 --col（列字母，如 D）")
            if args.value is None:
                _err("--mode set-col 需要 --value")
            ws = get_sheet(wb, args.sheet)
            n = op_set_col(ws, args.col, args.value, has_header=not args.no_header)
            extra_info = f"列 {args.col.upper()} 写入 {n} 个单元格"

        # ── add-sheet ─────────────────────────────────────────────────────────
        elif mode == "add-sheet":
            if not args.sheet_name:
                _err("--mode add-sheet 需要 --sheet-name")
            op_add_sheet(wb, args.sheet_name, args.sheet_position)
            extra_info = f"已新增工作表：{args.sheet_name}"

        # ── del-sheet ─────────────────────────────────────────────────────────
        elif mode == "del-sheet":
            if args.sheet is None:
                _err("--mode del-sheet 需要 --sheet（工作表名或索引）")
            name = op_del_sheet(wb, args.sheet)
            extra_info = f"已删除工作表：{name}"

        # ── rename-sheet ──────────────────────────────────────────────────────
        elif mode == "rename-sheet":
            if args.sheet is None:
                _err("--mode rename-sheet 需要 --sheet")
            if not args.sheet_name:
                _err("--mode rename-sheet 需要 --sheet-name（新名称）")
            old = op_rename_sheet(wb, args.sheet, args.sheet_name)
            extra_info = f"工作表重命名：{old} → {args.sheet_name}"

    except SystemExit:
        raise
    except Exception as e:
        _err(f"操作失败：{e}")

    # ── dry-run 输出 ──────────────────────────────────────────────────────────
    if args.dry_run:
        print(f"[write_excel] DRY-RUN：{mode} — {extra_info}（未写入）", file=sys.stderr)
        return

    # ── 备份 ─────────────────────────────────────────────────────────────────
    if not args.no_backup and os.path.exists(path):
        bak = backup(path)
        if bak:
            print(f"[write_excel] 备份：{bak}", file=sys.stderr)

    # ── 原子写入 ──────────────────────────────────────────────────────────────
    try:
        atomic_save(wb, path)
    except Exception as e:
        _err(f"保存失败：{e}")

    size = os.path.getsize(path)
    print(
        f"[write_excel] ✅ {mode} 成功 → {path}\n"
        f"              {extra_info}，文件大小：{size:,} 字节",
        file=sys.stderr,
    )


def _err(msg: str) -> None:
    print(f"[write_excel] 错误：{msg}", file=sys.stderr)
    sys.exit(1)


if __name__ == "__main__":
    main()
