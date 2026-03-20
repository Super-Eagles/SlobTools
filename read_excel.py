#!/usr/bin/env python3
"""
read_excel.py — 读取 Excel 文件（.xlsx / .xlsm），支持多种模式和输出格式。

用法示例：
  python read_excel.py <文件>                                  # 文件概览（info 模式）
  python read_excel.py <文件> --mode read                      # 读取全部数据
  python read_excel.py <文件> --mode read --sheet Sheet2       # 指定工作表
  python read_excel.py <文件> --mode read --start 5 --end 20   # 指定数据行范围
  python read_excel.py <文件> --mode read --cols 姓名 薪资     # 只读指定列
  python read_excel.py <文件> --mode read --limit 10           # 最多输出 10 行
  python read_excel.py <文件> --mode cell --cells A1 B3:C5     # 读取指定单元格
  python read_excel.py <文件> --mode formula --cells B2:D10    # 读取原始公式
  python read_excel.py <文件> --mode search --search "张三"    # 搜索
  python read_excel.py <文件> --mode stats                     # 数值列统计
  python read_excel.py <文件> --format json --out result.json  # JSON 输出到文件
"""

from __future__ import annotations
import sys
import os
import re
import json
import csv
import io
import argparse
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("错误：请先安装 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════════════════════

def _parse_single_ref(ref: str) -> tuple[int, int]:
    """解析单个单元格引用（如 A1），返回 (row, col)，均 1-indexed。"""
    ref = ref.strip().upper()
    m = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not m:
        _err(f"无效的单元格引用：{ref}")
    return int(m.group(2)), column_index_from_string(m.group(1))


def _parse_range(ref: str) -> list[tuple[int, int]]:
    """解析单元格引用，支持单格（A1）和范围（B3:C5）。"""
    ref = ref.strip().upper()
    if ":" in ref:
        a, b = ref.split(":", 1)
        sr, sc = _parse_single_ref(a)
        er, ec = _parse_single_ref(b)
        return [(r, c) for r in range(sr, er + 1) for c in range(sc, ec + 1)]
    return [_parse_single_ref(ref)]


def _get_sheet(wb, sheet_arg: str | None):
    """按名称或 0-based 索引获取工作表。"""
    if sheet_arg is None:
        return wb.active
    if sheet_arg in wb.sheetnames:
        return wb[sheet_arg]
    try:
        return wb.worksheets[int(sheet_arg)]
    except (ValueError, IndexError):
        pass
    _err(f"找不到工作表：'{sheet_arg}'，可用：{wb.sheetnames}")


def _read_sheet_data(ws) -> tuple[list[str], list[list[Any]]]:
    """读取工作表（首行为表头），返回 (header, data_rows)。"""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return [], []
    all_rows = []
    for r in range(1, max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        all_rows.append(row)
    if not all_rows:
        return [], []
    header = [str(h) if h not in (None, "") else f"列{i + 1}"
              for i, h in enumerate(all_rows[0])]
    return header, all_rows[1:]


# ══════════════════════════════════════════════════════════════════════════════
# 格式化输出
# ══════════════════════════════════════════════════════════════════════════════

def _display_width(s: str) -> int:
    """估算字符串终端显示宽度（CJK 字符按 2 计）。"""
    w = 0
    for c in s:
        cp = ord(c)
        if (0x4E00 <= cp <= 0x9FFF or 0x3400 <= cp <= 0x4DBF or
                0xFF00 <= cp <= 0xFFEF or 0x3000 <= cp <= 0x303F):
            w += 2
        else:
            w += 1
    return w


def _fmt_table(header: list[str], rows: list[list[Any]]) -> str:
    if not header:
        return "（无数据）"
    str_rows = [[str(v) if v is not None else "" for v in row] for row in rows]
    widths = [_display_width(h) for h in header]
    for row in str_rows:
        for i, cell in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], _display_width(cell))

    def pad(s: str, w: int) -> str:
        return s + " " * max(0, w - _display_width(s))

    sep_top = "┬".join("─" * w for w in widths)
    sep_mid = "┼".join("─" * w for w in widths)
    sep_bot = "┴".join("─" * w for w in widths)
    fmt_row = lambda cells: " │ ".join(pad(cells[i] if i < len(cells) else "", widths[i])
                                       for i in range(len(widths)))
    lines = [sep_top,
             fmt_row(header),
             sep_mid,
             *[fmt_row(r) for r in str_rows],
             sep_bot]
    return "\n".join(lines)


def _fmt_json(header: list[str], rows: list[list[Any]]) -> str:
    result = [{header[i]: (row[i] if i < len(row) else None)
               for i in range(len(header))} for row in rows]
    return json.dumps(result, ensure_ascii=False, indent=2, default=str)


def _fmt_csv(header: list[str], rows: list[list[Any]]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for row in rows:
        w.writerow([str(v) if v is not None else "" for v in row])
    return buf.getvalue()


def _fmt_raw(header: list[str], rows: list[list[Any]]) -> str:
    lines = ["\t".join(header)]
    for row in rows:
        lines.append("\t".join(str(v) if v is not None else "" for v in row))
    return "\n".join(lines)


def _format_output(header: list[str], rows: list[list[Any]], fmt: str) -> str:
    if fmt == "json":   return _fmt_json(header, rows)
    if fmt == "csv":    return _fmt_csv(header, rows)
    if fmt == "raw":    return _fmt_raw(header, rows)
    return _fmt_table(header, rows)


# ══════════════════════════════════════════════════════════════════════════════
# 各模式实现
# ══════════════════════════════════════════════════════════════════════════════

def mode_info(wb, args) -> str:
    lines = [f"文件：{args.filename}", f"工作表数：{len(wb.sheetnames)}", ""]
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        lines.append(f"  [{i}] {name}  （{max_row} 行 × {max_col} 列）")
        if max_row >= 1 and max_col >= 1:
            header = [str(ws.cell(row=1, column=c).value or f"列{c}")
                      for c in range(1, max_col + 1)]
            if len(header) > 20:
                lines.append(f"      列名（前20）：{header[:20]} …共{len(header)}列")
            else:
                lines.append(f"      列名：{header}")
        lines.append("")
    return "\n".join(lines).rstrip()


def mode_read(wb, args) -> str:
    ws = _get_sheet(wb, args.sheet)
    header, rows = _read_sheet_data(ws)
    if not header:
        return "（工作表为空）"

    # 列过滤
    if args.cols:
        indices = []
        for col_name in args.cols:
            if col_name in header:
                indices.append(header.index(col_name))
            else:
                try:
                    idx = column_index_from_string(col_name.upper()) - 1
                    if 0 <= idx < len(header):
                        indices.append(idx)
                    else:
                        print(f"警告：列 '{col_name}' 超出范围，已忽略", file=sys.stderr)
                except Exception:
                    print(f"警告：找不到列 '{col_name}'，已忽略", file=sys.stderr)
        if indices:
            header = [header[i] for i in indices]
            rows = [[row[i] if i < len(row) else "" for i in indices] for row in rows]

    total = len(rows)
    start = max(1, args.start) if args.start else 1
    end = min(total, args.end) if args.end else total
    sliced = rows[start - 1:end]
    if args.limit and args.limit > 0:
        sliced = sliced[:args.limit]

    result = _format_output(header, sliced, args.format)
    shown_end = start + len(sliced) - 1
    summary = (f"# 工作表：{ws.title}，数据共 {total} 行，"
               f"显示第 {start}～{shown_end} 行（不含表头）")
    return summary + "\n" + result


def mode_cell(wb, args) -> str:
    ws = _get_sheet(wb, args.sheet)
    if not args.cells:
        _err("--mode cell 需要 --cells 参数，例如：--cells A1 B3:C5")
    lines = [f"# 工作表：{ws.title}"]
    for ref in args.cells:
        for row, col in _parse_range(ref):
            val = ws.cell(row=row, column=col).value
            lines.append(f"  {get_column_letter(col)}{row} = {repr(val)}")
    return "\n".join(lines)


def mode_formula(wb, args) -> str:
    """重新以 data_only=False 加载，读取原始公式字符串。"""
    try:
        wb_raw = load_workbook(args.filename, data_only=False)
    except Exception as e:
        _err(f"加载文件失败（formula 模式）：{e}")
    ws = _get_sheet(wb_raw, args.sheet)
    if not args.cells:
        _err("--mode formula 需要 --cells 参数，例如：--cells B2:D10")
    lines = [f"# 工作表：{ws.title}（原始公式）"]
    for ref in args.cells:
        for row, col in _parse_range(ref):
            val = ws.cell(row=row, column=col).value
            ref_str = f"{get_column_letter(col)}{row}"
            if val is None:
                lines.append(f"  {ref_str} = （空）")
            elif isinstance(val, str) and val.startswith("="):
                lines.append(f"  {ref_str} = {val}")
            else:
                lines.append(f"  {ref_str} = {repr(val)}  （非公式）")
    return "\n".join(lines)


def mode_search(wb, args) -> str:
    ws = _get_sheet(wb, args.sheet)
    if not args.search:
        _err("--mode search 需要 --search 参数")
    header, rows = _read_sheet_data(ws)
    if not header:
        return "（工作表为空）"

    flags = 0 if args.case_sensitive else re.IGNORECASE
    if args.regex:
        try:
            compiled = re.compile(args.search, flags)
            match_fn = lambda s: compiled.search(str(s)) is not None
        except re.error as e:
            _err(f"正则表达式错误：{e}")
    else:
        needle = args.search if args.case_sensitive else args.search.lower()
        if args.case_sensitive:
            match_fn = lambda s: needle in str(s)
        else:
            match_fn = lambda s: needle in str(s).lower()

    # 确定限定列索引
    search_col_idx: int | None = None
    if args.col:
        if args.col in header:
            search_col_idx = header.index(args.col)
        else:
            try:
                search_col_idx = column_index_from_string(args.col.upper()) - 1
            except Exception:
                _err(f"找不到列：'{args.col}'")

    results = []
    for row_idx, row in enumerate(rows):
        if search_col_idx is not None:
            targets = [(search_col_idx, row[search_col_idx] if search_col_idx < len(row) else "")]
        else:
            targets = list(enumerate(row))
        if any(match_fn(v) for _, v in targets):
            results.append((row_idx + 2, row))  # +2 = 跳过表头，Excel 行从1计

    if not results:
        return f'未找到匹配项："{args.search}"'

    lines = [f"# 共找到 {len(results)} 行匹配（模式：{args.search!r}，工作表：{ws.title}）\n"]
    for excel_row, row_data in results:
        parts = []
        for ci, val in enumerate(row_data):
            col_ref = f"{get_column_letter(ci + 1)}{excel_row}"
            col_name = header[ci] if ci < len(header) else "?"
            parts.append(f"    {col_ref}[{col_name}]={repr(val)}")
        lines.append(f"  第 {excel_row} 行：")
        lines.extend(parts)
    return "\n".join(lines)


def mode_stats(wb, args) -> str:
    ws = _get_sheet(wb, args.sheet)
    header, rows = _read_sheet_data(ws)
    if not header:
        return "（工作表为空）"

    if args.cols:
        stat_indices = []
        for col_name in args.cols:
            if col_name in header:
                stat_indices.append(header.index(col_name))
            else:
                print(f"警告：找不到列 '{col_name}'，已忽略", file=sys.stderr)
    else:
        stat_indices = [ci for ci, _ in enumerate(header)
                        if any(isinstance(r[ci] if ci < len(r) else None, (int, float))
                               for r in rows)]

    if not stat_indices:
        return "（未找到数值列）"

    stat_header = ["列名", "非空行数", "最小值", "最大值", "平均值", "合计"]
    stat_rows = []
    for ci in stat_indices:
        vals = [r[ci] for r in rows if ci < len(r) and isinstance(r[ci], (int, float)) and r[ci] == r[ci]]
        if not vals:
            stat_rows.append([header[ci], 0, "—", "—", "—", "—"])
        else:
            s = sum(vals)
            stat_rows.append([header[ci], len(vals),
                               round(min(vals), 6), round(max(vals), 6),
                               round(s / len(vals), 6), round(s, 6)])

    return f"# 工作表：{ws.title}，数值统计\n" + _format_output(stat_header, stat_rows, args.format)


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

MODES   = ("info", "read", "cell", "formula", "search", "stats")
FORMATS = ("table", "json", "csv", "raw")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="读取 Excel 文件（.xlsx / .xlsm），支持多种模式和输出格式。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("filename", help="Excel 文件路径（.xlsx / .xlsm）")
    p.add_argument("--mode", choices=MODES, default="info",
                   help=f"操作模式（默认 info）：{' | '.join(MODES)}")
    p.add_argument("--sheet", default=None, metavar="NAME_OR_INDEX",
                   help="工作表名称或 0-based 索引（默认第一个）")
    # read
    p.add_argument("--start",  type=int, default=None, metavar="N",
                   help="[read] 起始数据行（1-indexed，不含表头）")
    p.add_argument("--end",    type=int, default=None, metavar="N",
                   help="[read] 结束数据行（1-indexed，不含表头）")
    p.add_argument("--cols",   nargs="+", default=None, metavar="COL",
                   help="[read/stats] 只处理指定列（列名或列字母，可多个）")
    p.add_argument("--limit",  type=int, default=None, metavar="N",
                   help="[read] 最多输出 N 行数据")
    # cell / formula
    p.add_argument("--cells",  nargs="+", default=None, metavar="REF",
                   help="[cell/formula] 单元格引用，如 A1 B3:C5")
    # search
    p.add_argument("--search", type=str, default=None, metavar="PATTERN",
                   help="[search] 搜索关键词或正则表达式")
    p.add_argument("--col",    type=str, default=None, metavar="COL",
                   help="[search] 限定搜索列（列名或列字母）")
    p.add_argument("--regex",  action="store_true",
                   help="[search] 将 --search 视为正则表达式")
    p.add_argument("--case-sensitive", action="store_true",
                   help="[search] 区分大小写（默认不区分）")
    # 输出
    p.add_argument("--format", choices=FORMATS, default="table",
                   help=f"输出格式（默认 table）：{' | '.join(FORMATS)}")
    p.add_argument("--out", type=str, default=None, metavar="FILE",
                   help="结果写入文件（UTF-8），默认输出到 stdout")
    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    path = os.path.abspath(args.filename)
    if not os.path.exists(path):
        _err(f"文件不存在：'{path}'")
    if not os.path.isfile(path):
        _err(f"'{path}' 不是普通文件")
    args.filename = path

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        _err(f"无法打开文件：{e}")

    dispatch = {
        "info":    mode_info,
        "read":    mode_read,
        "cell":    mode_cell,
        "formula": mode_formula,
        "search":  mode_search,
        "stats":   mode_stats,
    }
    try:
        result = dispatch[args.mode](wb, args)
    except SystemExit:
        raise
    except Exception as e:
        _err(f"执行失败：{e}")

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(result)
        print(f"结果已写入：{args.out}", file=sys.stderr)
    else:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        print(result)


def _err(msg: str) -> None:
    print(f"[read_excel] 错误：{msg}", file=sys.stderr)
    sys.exit(1)


if __name__ == "__main__":
    main()
